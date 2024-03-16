import time
from typing import Final
import openpyxl
from openpyxl import Workbook, load_workbook
import os
import excel2img
import pandas
from datetime import datetime
from telegram.ext import Application, CommandHandler, MessageHandler, filters
import telegram

def getir(hisse): #write dde links on excel
    wb = openpyxl.load_workbook('Template.xlsm', keep_vba=True)
    ws = wb.active
    ws["A1"] = hisse.upper()

    # Define the ranges for each set of data
    ranges = {
        "AEMIR_SAYISI": (3, 27),
        "AMIKTAR": (3, 27),
        "AFIYAT": (3, 27),
        "SFIYAT": (3, 27),
        "SMIKTAR": (3, 27),
        "SEMIR_SAYISI": (3, 27)
    }

    # Loop through each range and set the formulas in columns B to G
    for col_num, (key, (start, end)) in enumerate(ranges.items()):  # Start from column B
        for i in range(start, end + 1):
            ws[f"{chr(65 + col_num + 1)}{i}"].value = f"=MTX|DATA!{hisse.upper()}.{key}{i - 2}"

    # Save the workbook
    wb.save("Result.xlsm")
    wb.close()


# Defining the bot
TOKEN: Final = 'YOUR TOKEN HERE'
BOT_USERNAME: Final = '@YOUR BOT NAME'
allowed_chat_id = "" #if you want make the bot private

async def derinlik(update, context): #wait for a command

    hisse = " ".join(context.args)
    chat_id = update.message.chat_id
    if str(chat_id) == allowed_chat_id: #you should change this part if you dont want a private bot
        if hisse != "":
            await update.message.reply_text(hisse.upper() + " derinlik sorgulaması sıraya eklendi.")
            getir(hisse)

            time.sleep(2)
            os.system("start EXCEL.EXE Result.xlsm") #This part is essential, the excel file must be opened, saved and closed in order to update dde link values.
            time.sleep(18)
            os.system('taskkill /T /IM EXCEL.exe')
            time.sleep(3)

            excel_file = 'Result.xlsm'
            df = pandas.read_excel(excel_file, header=None, skiprows=[0, 1,27])

            for i in range(len(df)):
                for j in range(len(df.columns)):
                    var_name = f'{chr(65 + j)}{i + 3}'
                    globals()[var_name] = df.iloc[i, j]
                    if type(globals()[var_name]) == "str":
                        await update.message.reply_text("Bir sorun oluştu, girdiğiniz hissenin derinlik verisi artık okunamıyor olabilir. Lütfen geliştirici ile iletişime geçin.")
                    elif int(globals()[var_name]) == df.iloc[i, j]:
                        globals()[var_name] = int(globals()[var_name]) #pandas reads dde links as numbers


            wab = openpyxl.load_workbook('Template.xlsm', keep_vba=True)
            was = wab.active
            for i in range(len(df)):
                for j in range(len(df.columns)):
                    var_name = f'{chr(65 + j)}{i + 3}'
                    was[str(var_name)]=globals()[var_name]  #writing values as numbers
            now = datetime.today()
            was["A1"] = str(hisse.upper())
            was["B1"] = str(now.day) + "/" + str(now.month)+ "/" + str(now.year)
            was["F1"] = str(now.hour)+":"+str(now.minute)
            wab.save("son.xlsm")
            wab.close()

            excel2img.export_img("son.xlsm", "screenshot.png", "", "Sayfa1!A1:G28")


            await update.message.reply_photo("screenshot.png")

        else:
            await update.message.reply_text("Geçerli bir hisse giriniz...")
    else:
        await update.message.reply_text("Bu botu yalnızca xTicaret Vip üyeleri kullanabilir. Daha fazla bilgi için ziyaret et @xticaretdestek")
       


#starting

if __name__ == '__main__':
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler('derinlik', derinlik))

    print('Polling...')
    app.run_polling(poll_interval=3)





